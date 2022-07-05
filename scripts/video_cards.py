import csv
import re
import xlsxwriter
import openpyxl as openxl
import sys


def generate_headers(*values):
    return [{'header': value} for value in values]


def creates_workbook(csv_file, wb_path):
    headers = generate_headers('Placa de Vídeo', 'Preço', 'Loja', 'Modelo', 'Marca',
                               'Fornecedora', 'Memória', 'Tipo Memória', 'RTX',
                               'GTX', 'RX', 'GT', 'LHR', 'TI', 'Super', 'Overclock', 'Link')
    try:
        with open(csv_file, newline='') as f:
            reader = csv.reader(f)
            data = list(reader)
            column = openxl.utils.cell.get_column_letter(len(headers))
            row = len(data)
    except:
        sys.exit(f'Dados não encontrados em "{csv_file}"')

    try:
        wb = xlsxwriter.Workbook(wb_path)
        ws = wb.add_worksheet('videoCards')
        ws.add_table(f'A1:{column}{row}', {'data': data[1:-1], 'style':
                                           'Table Style Medium 5',
                                           'columns': headers})
        wb.close()
    except:
        sys.exit(f'Ocorreu um problema ao criar a planilha em "{wb_path}"')


def insert_data(i, worksheet, data):
    # workbook = openxl.load_workbook(wb_path)
    # worksheet = workbook[ws_name]

    # for i in range(2, worksheet.max_row+1):
    # data = build_data(i, worksheet)
    worksheet[f'A{i}'] = data['title']
    # worksheet[f'G{i}'] = worksheet[f'B{i}'].value
    # worksheet[f'B{i}'] = worksheet[f'D{i}'].value
    # worksheet[f'C{i}'] = worksheet[f'C{i}'].value
    worksheet[f'B{i}'] = data['price']
    worksheet[f'C{i}'] = data['store']
    worksheet[f'D{i}'] = data['model']
    worksheet[f'E{i}'] = data['brand']
    worksheet[f'F{i}'] = data['manufacturers']
    worksheet[f'G{i}'] = data['memory_size']
    worksheet[f'H{i}'] = data['memory_type']
    worksheet[f'I{i}'] = data['rtx']
    worksheet[f'J{i}'] = data['gtx']
    worksheet[f'K{i}'] = data['rx']
    worksheet[f'L{i}'] = data['gt']
    worksheet[f'M{i}'] = data['lhr']
    worksheet[f'N{i}'] = data['ti']
    worksheet[f'O{i}'] = data['super']
    worksheet[f'P{i}'] = data['oc']
    worksheet[f'Q{i}'] = data['link']
    # worksheet[f'J{i}'] = 'X' if 'RTX' in data['title'] else ''
    # worksheet[f'K{i}'] = 'X' if 'GTX' in data['title'] else ''
    # worksheet[f'L{i}'] = 'X' if 'RX' in data['title'] else ''
    # worksheet[f'M{i}'] = 'X' if 'GT' in data['title'] else ''
    # worksheet[f'N{i}'] = 'X' if 'LHR' in data['title'] else ''
    # worksheet[f'O{i}'] = 'X' if 'TI' in data['title'] else ''
    # worksheet[f'P{i}'] = 'X' if 'SUPER' in data['title'] else ''
    # worksheet[f'Q{i}'] = 'X' if 'OC' in data['title'] else ''

    # workbook.save(wb_path)


# def create_workbook(wb_path, ws_name='videoCards'):
#     try:
#         workbook = xlsxwriter.Workbook(wb_path)
#         workbook.add_worksheet(ws_name)
#         workbook.close()
#     except:
#         print(f'Ocorreu um erro ao criar a planilha "{wb_path}"')


# def clear(ws_name, wb_path):
#     workbook = openpyxl.load_workbook(wb_path)
#     worksheet = workbook[ws_name]

#     for row in range(1, worksheet.max_row+1):
#         for column in range(1, worksheet.max_column+1):
#             worksheet[f'{openpyxl.utils.cell.get_column_letter(column)}{row}'].style = 'Normal'
#             worksheet[f'{openpyxl.utils.cell.get_column_letter(column)}{row}'].value = None

#     workbook.save(wb_path)

# def get_data_from_neighbor_ws(ws_source, ws_dest, wb_path):
#     workbook = openxl.load_workbook(wb_path)
#     current_ws = workbook[ws_source]
#     final_ws = workbook[ws_dest]

#     count = final_ws.max_row
#     if count <= 1:
#         count += 1

#     for row in range(2, current_ws.max_row+1):
#         for column in range(1, current_ws.max_column+1):
#             final_ws[f'{openxl.utils.cell.get_column_letter(column)}{count}'] = current_ws[
#                 f'{openxl.utils.cell.get_column_letter(column)}{row}'].value
#         count += 1

#     workbook.save(wb_path)


# def create_header(ws_name, wb_path, **values):
#     wb = openxl.load_workbook(wb_path)
#     ws = wb[ws_name]

#     for key, value in values.items():
#         ws[f'{key.upper()}{1}'] = value

#     wb.save(wb_path)


# def load_csv_data(csv_file, wb_path, ws_name='videoCards'):
#     try:
#         # workbook = pd.ExcelFile(wb_path)
#         read_file = pd.read_csv(csv_file)
#         with pd.ExcelWriter(wb_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
#             read_file.to_excel(writer, sheet_name=ws_name, index=False, header=True)
#     except:
#         print(
#             f'Ocorreu um erro ao carregar os dados do arquivo "{os.path.basename(csv_file)}" na planilha "{os.path.basename(wb_path)}"')

# def build(workbook_path):
#     key_ws = 'finalResult'
#     try:
#         workbook = pd.ExcelFile(workbook_path)
#     except:
#         sys.exit(f'O caminho {workbook_path} informado não é válido')

#     workbook_name = os.path.basename(workbook_path)
#     sheets = workbook.sheet_names

#     if key_ws in workbook.sheet_names:
#         sheet_dest = key_ws
#     else:
#         sys.exit(f'A sheet{key_ws} não existe')

#     util.clear_worksheet(sheet_dest, workbook_name)

#     for sheet in sheets:
#         merge_all_csv(sheet, sheet_dest, workbook_path)

#     create_header(sheet_dest, workbook_path,
#                   a='Placa de Vídeo', b='Preço', c='Loja', d='Model', e='Marca',
#                   f='Fornecedora', g='Link', h='Memória', i='Tipo Memória', j='RTX',
#                   k='GTX', l='RX', m='GT', n='LHR', o='TI', p='Super', q='Overclock')

#     data_handle(sheet_dest, workbook_name)
